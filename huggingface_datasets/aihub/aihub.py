# Copyright 2021 hyeontae seo
# 
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
# 
#     http://www.apache.org/licenses/LICENSE-2.0
# 
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.


import os
import csv
import json
import copy
import hashlib
import glob
import functools
from re import A, X

import datasets
from openpyxl import load_workbook


_DESCRIPTION = """
Description is **formatted** as markdown.
It should also contain any processing which has been applied (if any),
(e.g. corrupted example skipped, images cropped,...):
"""

_CITATION = """
"""

_VERSION = datasets.Version('1.0.0', "")

_DATASET_ROOT = {
    'specialty_corpus' : 'AIHub/전문분야 말뭉치',
    'specialty_ko_en' : 'AIHub/전문분야한영',
    'korean_sns' : 'AIHub/한국어 SNS',
    'korean_dialog' : 'AIHub/한국어 대화',
    'korean_dialog_summary' : 'AIHub/한국어 대화 요약',
    'ko_en_trans_tech' : 'AIHub/한국어-영어 번역 말뭉치(기술과학)',
    'ko_en_trans_social' : 'AIHub/한국어-영어 번역 말뭉치(사회과학)',
    'ko_en_trans_parallel' : 'AIHub/한국어-영어 번역(병렬) 말뭉치',
    'ko_ja_trans' : 'AIHub/한국어-일본어 번역 말뭉치',
    'ko_zh_trans_tech' : 'AIHub/한국어-중국어 번역 말뭉치(기술과학)',
    'ko_zh_trans_social' : 'AIHub/한국어-중국어 번역 말뭉치(사회과학)',
}

 
_PAPER_SUMMARY_FEATURE = datasets.Features({
    'idx': datasets.Value("int32"),
    'doc_type': datasets.Value("string"),
    'doc_id': datasets.Value("string"),
    'title': datasets.Value("string"),
    'date': datasets.Value("string"),
    'reg_no': datasets.Value("string"),
    'ipc': datasets.Value("string"),
    'issued_by': datasets.Value("string"),
    'author': datasets.Value("string"),
    'summary_entire': datasets.Sequence({
        'orginal_text': datasets.Value("string"),
        'summary_text': datasets.Value("string"),
    }),
    'summary_section': datasets.Sequence({
        'orginal_text': datasets.Value("string"),
        'summary_text': datasets.Value("string"),
    }),
})

#specialty_corpus

_SPECIALITY_CORPUS_NE = datasets.Sequence({
    'id': datasets.Value("string"),
    'entity': datasets.Value("string"),
    'type': datasets.Value("string"),
    'begin': datasets.Value("int32"),
    'end': datasets.Value("int32"),
}),

_SPECIALITY_CORPUS_PAPER_FEATURE = datasets.Features({  # 전문분야 말뭉치 논문
    'idx': datasets.Value("int32"),
    'doc_type': datasets.Value("string"),
    'doc_id': datasets.Value("string"),
    'title': datasets.Value("string"),
    'date': datasets.Value("string"),
    'reg_no': datasets.Value("string"),
    'issued_by': datasets.Value("string"),
    'author': datasets.Value("string"),
    'ipc': datasets.Value("string"),
    'attr': datasets.Value("string"),
    'claim_no': datasets.Value("string"),
    'sentno': datasets.Value("string"),
    'text': datasets.Value("string"),
    'NE': datasets.Sequence({
        'id': datasets.Value("string"),
        'entity': datasets.Value("string"),
        'type': datasets.Value("string"),
        'begin': datasets.Value("int32"),
        'end': datasets.Value("int32"),
    }),
    #'NE': _SPECIALITY_CORPUS_NE,
})

_SPECIALITY_CORPUS_STATUTE_FEATURE = datasets.Features({  # 전문분야 말뭉치 법령
    'idx': datasets.Value("int32"),
    'doc_type': datasets.Value("string"),
    'doc_id': datasets.Value("string"),
    'title': datasets.Value("string"),
    'date': datasets.Value("string"),
    'reg_no': datasets.Value("string"),
    'sentno': datasets.Value("int32"),
    'sentence': datasets.Sequence({
        'attr': datasets.Value("string"),
        'text': datasets.Value("string"),
        'NE': datasets.Sequence({
            'id': datasets.Value("int32"),
            'entity': datasets.Value("string"),
            'type': datasets.Value("string"),
            'begin': datasets.Value("int32"),
            'end': datasets.Value("int32"),
        }),
    }),
})

_SPECIALITY_CORPUS_PATENT_N_FEATURE = datasets.Features({  # 전문분야 말뭉치 특허(숫자 파일)
    'idx': datasets.Value("int32"),
    'doc_type': datasets.Value("string"),
    'doc_id': datasets.Value("string"),
    'title': datasets.Value("string"),
    'date': datasets.Value("string"),
    'reg_no': datasets.Value("string"),
    'ipc': datasets.Value("string"),
    'attr': datasets.Value("string"),
    'sentno': datasets.Value("int32"),
    'claim_no': datasets.Value("string"),
    'sentence': datasets.Sequence({
        'text': datasets.Value("string"),
        'NE': datasets.Sequence({
            'id': datasets.Value("int32"),
            'entity': datasets.Value("string"),
            'type': datasets.Value("string"),
            'begin': datasets.Value("int32"),
            'end': datasets.Value("int32"),
        }),
    }),
})

_SPECIALITY_CORPUS_PATENT_A_FEATURE = datasets.Features({  # 전문분야 말뭉치 특허(z 파일)
    'idx': datasets.Value("int32"),
    'doc_type': datasets.Value("string"),
    'doc_id': datasets.Value("string"),
    'title': datasets.Value("string"),
    'date': datasets.Value("string"),
    'reg_no': datasets.Value("string"),
    'author': datasets.Value("string"),
    'ipc': datasets.Value("string"),
    'attr': datasets.Value("string"),
    'claim_no': datasets.Value("string"),
    'sentno': datasets.Value("string"),
    'text': datasets.Value("string"),
    'NE': datasets.Sequence({
            'id': datasets.Value("string"),
            'entity': datasets.Value("string"),
            'type': datasets.Value("string"),
            'begin': datasets.Value("int32"),
            'end': datasets.Value("int32"),
        }),
})

_SPECIALITY_CORPUS_LEADING_CASE_FEATURE = datasets.Features({  # 전문분야 말뭉치 판례
    'idx': datasets.Value("int32"),
    'doc_type': datasets.Value("string"),
    'doc_id': datasets.Value("string"),
    'title': datasets.Value("string"),
    'date': datasets.Value("string"),
    'reg_no': datasets.Value("string"),
    'issued_by': datasets.Value("string"),
    'sentno': datasets.Value("int32"),
    'sentence': datasets.Sequence({
        'text': datasets.Value("string"),
        'NE': datasets.Sequence({
            'id': datasets.Value("int32"),
            'entity': datasets.Value("string"),
            'type': datasets.Value("string"),
            'begin': datasets.Value("int32"),
            'end': datasets.Value("int32"),
        }),
    }),
})

_SPECIALITY_KO_EN_FEATURE = datasets.Features({ # 전문분야한영
    'idx': datasets.Value("int32"),
    'sid': datasets.Value("int32"),
    'domain': datasets.Value("string"),
    'korean': datasets.Value("string"),
    'english': datasets.Value("string"),
    'ko_num_of_phrases': datasets.Value("int32"),
    'en_num_of_words': datasets.Value("int32"),
    'length_classification': datasets.Value("int32"),
    'difficulty': datasets.Value("string"),
    'institution': datasets.Value("string"),
})

_KOREAN_SNS_FEATURE = datasets.Features({ # 한국어 SNS
    'idx': datasets.Value("int32"),
    'header': datasets.Features({
        'dialogueInfo': datasets.Features({
            'numberOfParticipants': datasets.Value("int32"),
            'numberOfUtterances': datasets.Value("int32"),
            'numberOfTurns': datasets.Value("int32"),
            'type': datasets.Value("string"),
            'topic': datasets.Value("string"),
            'dialogueID': datasets.Value("string"),
        }),
        'participantsInfo': datasets.Sequence({
            'age': datasets.Value("string"),
            'residentialProvince': datasets.Value("string"),
            'gender': datasets.Value("string"),
            'participantID': datasets.Value("string"),
        }),
    }),
    'body': datasets.Sequence({
        'utterance': datasets.Value("string"),
        'utteranceID': datasets.Value("string"),
        'participantID': datasets.Value("string"),
        'date': datasets.Value("string"),
        'turnID': datasets.Value("string"),
        'time': datasets.Value("string"),
    }),
})

_KOREAN_DIALOG_FEATURE = datasets.Features({    # 한국어 대화
    'idx': datasets.Value('int32'),
    'speaker': datasets.Value("string"),
    'sentence': datasets.Value("string"),
    'domain_id': datasets.Value("string"),
    'domain': datasets.Value("string"),
    'category': datasets.Value("string"),
    'speaker_id': datasets.Value("string"),
    'sentence_id': datasets.Value("string"),
    'main': datasets.Value("string"),
    'sub': datasets.Value("string"),
    'qa': datasets.Value("string"),
    'qacnct': datasets.Value("string"),
    'mq': datasets.Value("string"),
    'sq': datasets.Value("string"),
    'ua': datasets.Value("string"),
    'sa': datasets.Value("string"),
    'object_name': datasets.Value("string"),
    'glossary': datasets.Value("string"),
    'knowledge_base': datasets.Value("string"),
})

_KOREAN_DIALOG_SUMMARY_FEATURE = datasets.Features({    # 한국어 대화 요약
    'idx': datasets.Value("int32"),
    'header': datasets.Features({
        'dialogueInfo': datasets.Features({
            'numberOfParticipants': datasets.Value("int32"),
            'numberOfUtterances': datasets.Value("int32"),
            'numberOfTurns': datasets.Value("int32"),
            'type': datasets.Value("string"),
            'topic': datasets.Value("string"),
            'dialogueID': datasets.Value("string"),
        }),
        'participantsInfo': datasets.Sequence({
            'age': datasets.Value("string"),
            'residentialProvince': datasets.Value("string"),
            'gender': datasets.Value("string"),
            'participantID': datasets.Value("string"),
        }),
    }),
    'body': datasets.Features({
        'dialogue': datasets.Sequence({
            'utterance': datasets.Value("string"),
            'utteranceID': datasets.Value("string"),
            'participantID': datasets.Value("string"),
            'date': datasets.Value("string"),
            'turnID': datasets.Value("string"),
            'time': datasets.Value("string"),
        }),
        'summary': datasets.Value("string"),
    }),
})

_TRANSLATION_KO_EN_FEATURE = datasets.Features({    # 한국어-영어 번역 말뭉치(기술과학)/(사회과학)
    'idx': datasets.Value("int32"),
    'sn': datasets.Value("string"),
    'file_name': datasets.Value("string"),
    'data_set': datasets.Value("string"),
    'domain': datasets.Value("string"),
    'subdomain': datasets.Value("string"),
    'source': datasets.Value("string"),
    'ko': datasets.Value("string"),
    'mt': datasets.Value("string"),
    'en': datasets.Value("string"),
    'source_language': datasets.Value("string"),
    'target_language': datasets.Value("string"),
    'license': datasets.Value("string"),
    'style': datasets.Value("string"),
})

_TRANSLATION_KO_EN_PARALLEL_INFORMAL = datasets.Features({  # 한국어-영어 번역(병렬) 말뭉치 구어체
    'idx': datasets.Value("int32"),
    'sid': datasets.Value("int32"),
    'original': datasets.Value("string"),
    'translated': datasets.Value("string"),
})

_TRANSLATION_KO_EN_PARALLEL_CONVERSATIONAL = datasets.Features({    # 한국어-영어 번역(병렬) 말뭉치 대화체
    'idx': datasets.Value("int32"),
    'main_category': datasets.Value("string"),
    'sub_category': datasets.Value("string"),
    'situation': datasets.Value("string"),
    'setNr': datasets.Value("int32"),
    'talker': datasets.Value("string"),
    'original': datasets.Value("string"),
    'translated': datasets.Value("string"),
})

_TRANSLATION_KO_EN_PARALLEL_NEWS = datasets.Features({  # 한국어-영어 번역(병렬) 말뭉치 문어체_뉴스
    'idx': datasets.Value("int32"),
    'id': datasets.Value("int32"),
    'date': datasets.Value("string"),
    'auto_class1': datasets.Value("string"),
    'auto_class2': datasets.Value("string"),
    'auto_class3': datasets.Value("string"),
    'url': datasets.Value("string"),
    'press': datasets.Value("string"),
    'original': datasets.Value("string"),
    'translated': datasets.Value("string"),
})

_TRANSLATION_KO_EN_PARALLEL_CULTURE = datasets.Features({   # 한국어-영어 번역(병렬) 말뭉치 문어체_한국문화
    'idx': datasets.Value("int32"),
    'id': datasets.Value("float"),
    'keyword': datasets.Value("string"),
    'original': datasets.Value("string"),
    'translated': datasets.Value("string"),
})

_TRANSLATION_KO_EN_PARALLEL_ORDINANCE_WEB = datasets.Features({ # 한국어-영어 번역(병렬) 말뭉치 문어체_조례 / 지자체 웹사이트
    'idx': datasets.Value("int32"),
    'id': datasets.Value("float"),
    'local_government': datasets.Value("string"),
    'original': datasets.Value("string"),
    'translated': datasets.Value("string"),
})

_TRANSLATION_KO_JA_FEATURE = datasets.Features({    # 한국어-일본어 번역 말뭉치
    'idx': datasets.Value("int32"),
    'sid': datasets.Value("string"),
    'domain': datasets.Value("string"),
    'korean': datasets.Value("string"),
    'japanese': datasets.Value("string"),
    'ko_num_of_phrases': datasets.Value("int32"),
    'ja_num_of_words': datasets.Value("int32"),
    'length_classification': datasets.Value("int32"),
    'source': datasets.Value("string"),
    'institution': datasets.Value("string"),
})

_TRANSLATION_KO_ZH_FEATURE = datasets.Features({    # 한국어-중국어 번역 말뭉치(기술과학)/(사회과학)
    'idx': datasets.Value("int32"),
    'sid': datasets.Value("string"),
    'domain': datasets.Value("string"),
    'korean': datasets.Value("string"),
    'chinese': datasets.Value("string"),
    'ko_num_of_phrases': datasets.Value("int32"),
    'zh_num_of_words': datasets.Value("int32"),
    'length_classification': datasets.Value("int32"),
    'source': datasets.Value("string"),
    'institution': datasets.Value("string"),
})

def _NE_list(data_list):
    result = list()
    for data in data_list:
        _id = data['id']
        _entity = data['entity']
        _type = data['type']
        _begin = data['begin']
        _end = data['end']
        result.append({
            "id": _id,
            "entity": _entity,
            "type": _type,
            "begin": _begin,
            "end": _end
        })
    return result

def _parsing_specialty_corpus_paper(file_path): # 전문분야 말뭉치 논문

    with open(file_path, mode='r') as f:
        obj = json.load(f)
        idx = 0
        for idx, doc in enumerate(obj['data']):
            _idx = idx
            _doc_type = doc['doc_type']
            _doc_id = doc['doc_id']
            _title = doc['title']
            _date = doc['date']
            _reg_no = doc['reg_no']
            _issued_by = doc['issued_by']
            _author = doc['author']
            _ipc = doc['ipc']
            _attr = doc['attr']
            _claim_no = doc['claim_no']
            _sentno = doc['sentno']
            _text = doc['text']
            _NE = _NE_list(doc['NE'])
            
            yield _idx, {
                'idx': _idx,
                'doc_type': _doc_type,
                'doc_id': _doc_id,
                'title': _title,
                'date': _date,
                'reg_no': _reg_no,
                'issued_by': _issued_by,
                'author': _author,
                'ipc': _ipc,
                'attr': _attr,
                'claim_no': _claim_no,
                'sentno': _sentno,
                'text': _text,
                'NE': _NE,
            }

def _statute_sentence_list(data_list):
    result = list()
    for data in data_list:
        _attr = data['attr']
        _text = data['text']
        _NE = _NE_list(data['NE'])
        result.append({
            "attr": _attr,
            "text": _text,
            "NE": _NE,
        })
    return result

def _parsing_specialty_corpus_statute(file_path):   # 전문분야 말뭉치 법령
    with open(file_path, mode='r') as f:
        obj = json.load(f)
        
        for idx, doc in enumerate(obj['data']):
            _idx = idx
            _doc_type = doc['doc_type']
            _doc_id = doc['doc_id']
            _title = doc['title']
            _date = doc['date']
            _reg_no = doc['reg_no']
            _sentno = doc['sentno']
            _sentence = _statute_sentence_list(doc['sentence'])

            yield _idx, {
                'idx': _idx,
                'doc_type': _doc_type,
                'doc_id': _doc_id,
                'title': _title,
                'date': _date,
                'reg_no': _reg_no,
                'sentno': _sentno,
                'sentence': _sentence,
            }

def _patent_n_sentence_list(data_list):
    result = list()
    for data in data_list:
        _text = data['text']
        _NE = _NE_list(data['NE'])
        result.append({
            "text": _text,
            "NE": _NE,
        })
    return result

def _parsing_specialty_corpus_patent_n(file_path):  # 전문분야 말뭉치 특허(숫자 파일)
    with open(file_path, mode='r') as f:
        obj = json.load(f)
        
        for idx, doc in enumerate(obj['data']):
            _idx = idx
            _doc_type = doc['doc_type']
            _doc_id = doc['doc_id']
            _title = doc['title']
            _date = doc['date']
            _reg_no = doc['reg_no']
            _ipc = doc['ipc']
            _attr = doc['attr']
            _sentno = doc['sentno']
            if 'claim_no' in doc:
                _claim_no = doc['claim_no']
            else:
                _claim_no = ''
            _sentence = _patent_n_sentence_list(doc['sentence'])

            yield _idx, {
                'idx': _idx,
                'doc_type': _doc_type,
                'doc_id': _doc_id,
                'title': _title,
                'date': _date,
                'reg_no': _reg_no,
                'ipc' : _ipc,
                "attr": _attr,
                'sentno': _sentno,
                'claim_no': _claim_no,
                'sentence': _sentence,
            }

def _parsing_specialty_corpus_patent_a(file_path):  # 전문분야 말뭉치 특허(z 파일)
    with open(file_path, mode='r') as f:
        obj = json.load(f)

        for idx, doc in enumerate(obj['data']):
            _idx = idx
            _doc_type = doc['doc_type']
            _doc_id = doc['doc_id']
            _title = doc['title']
            _date = doc['date']
            _reg_no = doc['reg_no']
            _author = doc['author']
            _ipc = doc['ipc']
            _attr = doc['attr']
            _claim_no = doc['claim_no']
            _sentno = doc['sentno']
            _text = doc['text']
            _NE = _NE_list(doc['NE'])

            yield _idx, {
                'idx': _idx,
                'doc_type': _doc_type,
                'doc_id': _doc_id,
                'title': _title,
                'date': _date,
                'reg_no': _reg_no,
                'author': _author,
                'ipc': _ipc,
                'attr': _attr,
                'claim_no': _claim_no,
                'sentno': _sentno,
                'text': _text,
                'NE': _NE,
            }

def _leading_case_sentence_list(data_list):
    result = list()
    for data in data_list:
        _text = data['text']
        _NE = _NE_list(data['NE'])
        result.append({
            "text": _text,
            "NE": _NE,
        })
    return result

def _parsing_specialty_corpus_leading_case(file_path):  # 전문분야 말뭉치 판례
    with open(file_path, mode='r') as f:
        obj = json.load(f)

        for idx, doc in enumerate(obj['data']):
            _idx = idx
            _doc_type = doc['doc_type']
            _doc_id = doc['doc_id']
            _title = doc['title']
            _date = doc['date']
            _reg_no = doc['reg_no']
            _issued_by = doc['issued_by']
            _sentno = doc['sentno']
            _sentence = _leading_case_sentence_list(doc['sentence'])

            yield _idx, {
                'idx': -idx,
                'doc_type': _doc_type,
                'doc_id': _doc_id,
                'title': _title,
                'date': _date,
                'reg_no': _reg_no,
                'issued_by' : _issued_by,
                'sentno': _sentno,
                'sentence': _sentence,
            }

def _parsing_specialty_ko_en(file_path):    # 전문분야한영
    with open(file_path, mode='r') as f:
        obj = json.load(f)

        for idx, doc in enumerate(obj):
            _idx = idx
            _sid = doc['sid']
            _domain = doc['분야']
            _korean = doc['한국어']
            _english = doc['영어']
            _ko_num_of_phrases = doc['한국어_어절수']
            _en_num_of_words = doc['영어_단어수']
            _length_classification = doc['길이_분류']
            _difficulty = doc['난이도']
            _institution = doc['수행기관']
            
            yield _idx, {
                'idx': _idx,
                'sid': _sid,
                'domain': _domain,
                'korean': _korean,
                'english': _english,
                'ko_num_of_phrases': _ko_num_of_phrases,
                'en_num_of_words': _en_num_of_words,
                'length_classification': _length_classification,
                'difficulty': _difficulty,
                'institution': _institution,
            }

def _parsing_korean_sns(file_path): # 한국어 SNS
    with open(file_path, mode='r') as f:
        obj = json.load(f)

        for idx, doc in enumerate(obj['data']):
            _idx = idx
            _header = doc['header']
            _body = doc['body']
            
            yield _idx, {
                'idx': _idx,
                'header': _header,
                'body': _body,
            }

def _parsing_korean_dialog(file_path):  # 한국어 대화
    load_wb = load_workbook(file_path, data_only=True)
    load_ws = load_wb.active

    for idx, row in enumerate(load_ws.rows):
        _idx = idx

        _speaker = str(load_ws.cell(idx+2, 1).value)
        _sentence = str(load_ws.cell(idx+2, 2).value)
        _domain_id = str(load_ws.cell(idx+2, 3).value)
        _domain = str(load_ws.cell(idx+2, 4).value)
        _category = str(load_ws.cell(idx+2, 5).value)
        _speaker_id = str(load_ws.cell(idx+2, 6).value)
        _sentence_id = str(load_ws.cell(idx+2, 7).value)
        _main = str(load_ws.cell(idx+2, 8).value)
        if load_ws.cell(idx+2, 9).value is not None:
            _sub = str(load_ws.cell(idx+2, 9).value)
        else:
            _sub = ''
        _qa = str(load_ws.cell(idx+2, 10).value)
        if load_ws.cell(idx+2, 11).value is not None:
            _qacnct = str(load_ws.cell(idx+2, 11).value)
        else:
            _qacnct = ''
        if load_ws.cell(idx+2, 12).value is not None:
            _mq = str(load_ws.cell(idx+2, 12).value)
        else:
            _mq = ''
        if load_ws.cell(idx+2, 13).value is not None:
            _sq = str(load_ws.cell(idx+2, 13).value)
        else:
            _sq = ''
        if load_ws.cell(idx+2, 14).value is not None:
            _ua = str(load_ws.cell(idx+2, 14).value)
        else:
            _ua = ''
        if load_ws.cell(idx+2, 15).value is not None:
            _sa = str(load_ws.cell(idx+2, 15).value)
        else:
            _sa = ''
        if load_ws.cell(idx+2, 16).value is not None:
            _object_name = str(load_ws.cell(idx+2, 16).value)
        else:
            _object_name = ''
        if load_ws.cell(idx+2, 17).value is not None:
            _glossary = str(load_ws.cell(idx+2, 17).value)
        else:
            _glossary = ''
        if load_ws.cell(idx+2, 18).value is not None:
            _knowledge_base = str(load_ws.cell(idx+2, 18).value)
        else:
            _knowledge_base = ''
        
        yield _idx, {
            'idx': _idx,
            'speaker': _speaker,
            'sentence': _sentence,
            'domain_id': _domain_id,
            'domain': _domain,
            'category': _category,
            'speaker_id': _speaker_id,
            'sentence_id': _sentence_id,
            'main': _main,
            'sub': _sub,
            'qa': _qa,
            'qacnct': _qacnct,
            'mq': _mq,
            'sq': _sq,
            'ua': _ua,
            'sa': _sa,
            'object_name': _object_name,
            'glossary': _glossary,
            'knowledge_base': _knowledge_base,
        }

def _parsing_korean_dialog_summary(file_path):  # 한국어 대화 요약
    with open(file_path, mode='r') as f:
        obj = json.load(f)

        for idx, doc in enumerate(obj['data']):
            _idx = idx
            _header = doc['header']
            _body = doc['body']
            
            yield _idx, {
                'idx': _idx,
                'header': _header,
                'body': _body,
            }

def _parsing_translation_ko_en(file_path):  # 한국어-영어 번역 말뭉치(기술과학)/(사회과학)
    with open(file_path, mode='r') as f:
        obj = json.load(f)

        for idx, doc in enumerate(obj['data']):
            _idx = idx
            _sn = doc['sn']
            _file_name = doc['file_name']
            _data_set = doc['data_set']
            _domain = doc['domain']
            _subdomain = doc['subdomain']
            _source = doc['source']
            _ko = doc['ko']
            _mt = doc['mt']
            _en = doc['en']
            _source_language = doc['source_language']
            _target_language = doc['target_language']
            _license = doc['license']
            _style = doc['style']

            yield _idx, {
                'idx': _idx,
                'sn': _sn,
                'file_name': _file_name,
                'data_set': _data_set,
                'domain': _domain,
                'subdomain': _subdomain,
                'source': _source,
                'ko': _ko,
                'mt': _mt,
                'en': _en,
                'source_language': _source_language,
                'target_language': _target_language,
                'license': _license,
                'style': _style,
            }

def _parsing_ko_en_parallel_informal(file_path):    # 한국어-영어 번역(병렬) 말뭉치 구어체
    load_wb = load_workbook(file_path, data_only=True)
    load_ws = load_wb.active
    row_idx = 2
    for idx, _ in enumerate(load_ws.rows):
        if isinstance(load_ws.cell(row_idx+idx, 1).value, int):
            _idx = idx
            _sid = load_ws.cell(row_idx+idx, 1).value
            _original = load_ws.cell(row_idx+idx, 2).value
            _translated = load_ws.cell(row_idx+idx, 3).value
            yield _idx, {
                'idx': _idx,
                'sid': _sid,
                'original': _original,
                'translated': _translated,
            }
        else:
            row_idx -= 1

def _parsing_ko_en_parallel_conversational(file_path):  # 한국어-영어 번역(병렬) 말뭉치 대화체
    load_wb = load_workbook(file_path, data_only=True)
    load_ws = load_wb.active
    row_idx = 2
    for idx, _ in enumerate(load_ws.rows):
        if isinstance(load_ws.cell(row_idx+idx, 4).value, int):
            _idx = idx
            _main_category = load_ws.cell(row_idx+idx, 1).value
            _sub_category = load_ws.cell(row_idx+idx, 2).value
            _situation = load_ws.cell(row_idx+idx, 3).value
            _setNr = load_ws.cell(row_idx+idx, 4).value
            _talker = load_ws.cell(row_idx+idx, 5).value
            _original = load_ws.cell(row_idx+idx, 6).value
            _translated = load_ws.cell(row_idx+idx, 7).value


            yield _idx, {
                'idx': _idx,
                'main_category': _main_category,
                'sub_category': _sub_category,
                'situation': _situation,
                'setNr': _setNr,
                'talker': _talker,
                'original': _original,
                'translated': _translated,
            }
        else:
            row_idx -= 1

def _parsing_ko_en_parallel_news(file_path):    # 한국어-영어 번역(병렬) 말뭉치 문어체_뉴스
    load_wb = load_workbook(file_path, data_only=True)
    load_ws = load_wb.active
    row_idx = 2
    for idx, _ in enumerate(load_ws.rows):
        if isinstance(load_ws.cell(row_idx+idx, 1).value, int):
            _idx = idx
            _id = load_ws.cell(row_idx+idx, 1).value
            _date = str(load_ws.cell(row_idx+idx, 2).value)
            _auto_class1 = load_ws.cell(row_idx+idx, 3).value
            _auto_class2 = load_ws.cell(row_idx+idx, 4).value
            _auto_class3 = load_ws.cell(row_idx+idx, 5).value
            _url = load_ws.cell(row_idx+idx, 6).value
            _press = load_ws.cell(row_idx+idx, 7).value
            _original = load_ws.cell(row_idx+idx, 8).value
            _translated = load_ws.cell(row_idx+idx, 9).value
            yield _idx, {
                'idx': _idx,
                'id': _id,
                'date': _date,
                'auto_class1': _auto_class1,
                'auto_class2': _auto_class2,
                'auto_class3': _auto_class3,
                'url': _url,
                'press': _press,
                'original': _original,
                'translated': _translated,
            }
        else:
            row_idx -= 1

def _parsing_ko_en_parallel_culture(file_path): # 한국어-영어 번역(병렬) 말뭉치 문어체_한국문화
    load_wb = load_workbook(file_path, data_only=True)
    load_ws = load_wb.active
    row_idx = 2
    for idx, _ in enumerate(load_ws.rows):
        if isinstance(load_ws.cell(row_idx+idx, 1).value, float):
            _idx = idx
            _id = load_ws.cell(row_idx+idx, 1).value
            _keyword = load_ws.cell(row_idx+idx, 2).value
            _original = load_ws.cell(row_idx+idx, 3).value
            _translated = load_ws.cell(row_idx+idx, 4).value
            
            yield _idx, {
                'idx': _idx,
                'id': _id,
                'keyword': _keyword,
                'original': _original,
                'translated': _translated,
            }
        else:
            row_idx -= 1

def _parsing_ko_en_parallel_ordinance_web(file_path):   # 한국어-영어 번역(병렬) 말뭉치 문어체_조례 / 지자체 웹사이트
    load_wb = load_workbook(file_path, data_only=True)
    load_ws = load_wb.active
    row_idx = 2
    for idx, _ in enumerate(load_ws.rows):
        if isinstance(load_ws.cell(row_idx+idx, 1).value, float):
            _idx = idx
            _id = load_ws.cell(row_idx+idx, 1).value
            _local_government = load_ws.cell(row_idx+idx, 2).value
            _original = load_ws.cell(row_idx+idx, 3).value
            _translated = load_ws.cell(row_idx+idx, 4).value

            yield _idx, {
                'idx': _idx,
                'id': _id,
                'local_government': _local_government,
                'original': _original,
                'translated': _translated,
            }
        else:
            row_idx -= 1

def _parsing_specialty_ko_ja(file_path):    # 한국어-일본어 번역 말뭉치
    with open(file_path, mode='r') as f:
        obj = json.load(f)

        for idx, doc in enumerate(obj):
            _idx = idx
            _sid = doc['관리번호']
            _domain = doc['분야']
            _korean = doc['한국어']
            _japanese = doc['일본어']
            _ko_num_of_phrases = doc['한국어_어절수']
            _ja_num_of_words = doc['일본어_글자수']
            _length_classification = doc['길이_분류']
            _source = doc['출처']
            _institution = doc['수행기관']
            
            yield _idx, {
                'idx': _idx,
                'sid': _sid,
                'domain': _domain,
                'korean': _korean,
                'japanese': _japanese,
                'ko_num_of_phrases': _ko_num_of_phrases,
                'ja_num_of_words': _ja_num_of_words,
                'length_classification': _length_classification,
                'source': _source,
                'institution': _institution,
            }

def _parsing_specialty_ko_zh(file_path):    # 한국어-중국어 번역 말뭉치(기술과학)/(사회과학)
    with open(file_path, mode='r') as f:
        obj = json.load(f)

        for idx, doc in enumerate(obj):
            _idx = idx
            _sid = doc['관리번호']
            _domain = doc['분야']
            _korean = doc['한국어']
            _chinese = doc['중국어']
            _ko_num_of_phrases = doc['한국어_어절수']
            _zh_num_of_words = doc['중국어_글자수']
            _length_classification = doc['길이_분류']
            _source = doc['출처']
            _institution = doc['수행기관']
            
            yield _idx, {
                'idx': _idx,
                'sid': _sid,
                'domain': _domain,
                'korean': _korean,
                'chinese': _chinese,
                'ko_num_of_phrases': _ko_num_of_phrases,
                'zh_num_of_words': _zh_num_of_words,
                'length_classification': _length_classification,
                'source': _source,
                'institution': _institution,
            }

def _hash_text(text):
    return hashlib.md5(text.encode("utf-8")).hexdigest()

def _filter_fn_hash_id(uid, split_fn):
    hash_id = _hash_text(str(uid))
    val = int(hash_id, 16)
    return split_fn(val)

_DEFAULT_RAW_CORPUS_SPLIT = {
              'source': [datasets.Split.TRAIN],
              'split': {
                datasets.Split.TRAIN: lambda x: x % 1000 > 0,
                datasets.Split.VALIDATION: lambda x: x % 1000 == 0,
              }}

_DEFAULT_DOWNSTREAMTASK_CORPUS_SPLIT = {
              'source': [datasets.Split.TRAIN],
              'split': {
                datasets.Split.TRAIN: lambda x: x % 10 > 1,
                datasets.Split.VALIDATION: lambda x: x % 10 == 0,
                datasets.Split.TEST: lambda x: x % 10 == 1,
              }}

class AIHubConfig(datasets.BuilderConfig):
    def __init__(self,
                 name,
                 data_root,
                 feature,
                 data_sp_path,
                 reading_fn,
                 parsing_fn,
                 additional_data_root=None,
                 homepage='https://aihub.or.kr/',
                 split_fn=None,
                 metadata=None,
                 **kwargs):
        super(AIHubConfig, self).__init__(
            name=name,
            version=_VERSION,
            **kwargs
        )
        self.data_root = data_root
        self.feature = feature
        self.data_sp_path = data_sp_path
        self.reading_fn = reading_fn
        self.parsing_fn = parsing_fn
        self.additional_data_root = additional_data_root
        self.homepage = homepage
        self.split_fn = split_fn
        self.metadata = metadata

class AIHub(datasets.GeneratorBasedBuilder):
    """DatasetBuilder for AIHub dataset."""

    RELEASE_NOTES = {
        '1.0.0': 'Initial release.',
    }

    BUILDER_CONFIGS = [
        AIHubConfig(
            name='specialty_corpus.paper.(v1.0)',
            data_root=_DATASET_ROOT['specialty_corpus'],
            feature=_SPECIALITY_CORPUS_PAPER_FEATURE,
            data_sp_path={datasets.Split.TRAIN: ['Training/논문*.json'],
                          datasets.Split.VALIDATION: ['Validation/논문*.json']},
            reading_fn=_parsing_specialty_corpus_paper,
            parsing_fn=lambda x:x,
        ),

        AIHubConfig(
            name='specialty_corpus.statute.(v1.0)',
            data_root=_DATASET_ROOT['specialty_corpus'],
            feature=_SPECIALITY_CORPUS_STATUTE_FEATURE,
            data_sp_path={datasets.Split.TRAIN: ['Training/법령*.json'],
                          datasets.Split.VALIDATION: ['Validation/법령*.json']},
            reading_fn=_parsing_specialty_corpus_statute,
            parsing_fn=lambda x:x,
        ),        
        
        AIHubConfig(
            name='specialty_corpus.patent_n.(v1.0)',
            data_root=_DATASET_ROOT['specialty_corpus'],
            feature=_SPECIALITY_CORPUS_PATENT_N_FEATURE,
            data_sp_path={datasets.Split.TRAIN: [r'Training/특허_[0-9][0-9].json'],
                          datasets.Split.VALIDATION: [r'Training/특허_[0-9][0-9].json']},
            reading_fn=_parsing_specialty_corpus_patent_n,
            parsing_fn=lambda x:x,
        ),

        AIHubConfig(
            name='specialty_corpus.patent_a.(v1.0)',
            data_root=_DATASET_ROOT['specialty_corpus'],
            feature=_SPECIALITY_CORPUS_PATENT_A_FEATURE,
            data_sp_path={datasets.Split.TRAIN: ['Training/특허_z*.json'],
                          datasets.Split.VALIDATION: ['Training/특허_z*.json']},
            reading_fn=_parsing_specialty_corpus_patent_a,
            parsing_fn=lambda x:x,
            split_fn=_DEFAULT_RAW_CORPUS_SPLIT,
        ),

        AIHubConfig(
            name='specialty_corpus.leading_case.(v1.0)',
            data_root=_DATASET_ROOT['specialty_corpus'],
            feature=_SPECIALITY_CORPUS_LEADING_CASE_FEATURE,
            data_sp_path={datasets.Split.TRAIN: ['Training/판례*.json'],
                          datasets.Split.VALIDATION: ['Validation/판례*.json']},
            reading_fn=_parsing_specialty_corpus_leading_case,
            parsing_fn=lambda x:x,
        ),

        AIHubConfig(
            name='specialty_ko_en.(v1.0)',
            data_root=_DATASET_ROOT['specialty_ko_en'],
            feature=_SPECIALITY_KO_EN_FEATURE,
            data_sp_path={datasets.Split.TRAIN: ['Training/*.json'],
                          datasets.Split.VALIDATION: ['Validation/*.json']},
            reading_fn=_parsing_specialty_ko_en,
            parsing_fn=lambda x:x,
        ),

        AIHubConfig(
            name='korean_sns.(v1.0)',
            data_root=_DATASET_ROOT['korean_sns'],
            feature=_KOREAN_SNS_FEATURE,
            data_sp_path={datasets.Split.TRAIN: ['Training/*.json'],
                          datasets.Split.VALIDATION: ['Training/*.json']},
            reading_fn=_parsing_korean_sns,
            parsing_fn=lambda x:x,
            split_fn=_DEFAULT_RAW_CORPUS_SPLIT,
        ),

        AIHubConfig(
            name='korean_dialog.(v1.0)',
            data_root=_DATASET_ROOT['korean_dialog'],
            feature=_KOREAN_DIALOG_FEATURE,
            data_sp_path={datasets.Split.TRAIN: ['*.xlsx'],
                          datasets.Split.VALIDATION: ['*.xlsx']},
            reading_fn=_parsing_korean_dialog,
            parsing_fn=lambda x:x,
            split_fn=_DEFAULT_RAW_CORPUS_SPLIT,
        ),

        AIHubConfig(
            name='korean_dialog_summary.(v1.0)',
            data_root=_DATASET_ROOT['korean_dialog_summary'],
            feature=_KOREAN_DIALOG_SUMMARY_FEATURE,
            data_sp_path={datasets.Split.TRAIN: ['Training/*.json'],
                          datasets.Split.VALIDATION: ['Validation/*.json']},
            reading_fn=_parsing_korean_dialog_summary,
            parsing_fn=lambda x:x,
        ),

        AIHubConfig(
            name='translation_ko_en_tech.(v1.0)',
            data_root=_DATASET_ROOT['ko_en_trans_tech'],
            feature=_TRANSLATION_KO_EN_FEATURE,
            data_sp_path={datasets.Split.TRAIN: ['Training/*.json'],
                          datasets.Split.VALIDATION: ['Validation/*.json']},
            reading_fn=_parsing_translation_ko_en,
            parsing_fn=lambda x:x,
        ),

        AIHubConfig(
            name='translation_ko_en_social.(v1.0)',
            data_root=_DATASET_ROOT['ko_en_trans_social'],
            feature=_TRANSLATION_KO_EN_FEATURE,
            data_sp_path={datasets.Split.TRAIN: ['Training/*.json'],
                          datasets.Split.VALIDATION: ['Validation/*.json']},
            reading_fn=_parsing_translation_ko_en,
            parsing_fn=lambda x:x,
        ),

        AIHubConfig(
            name='ko_en_trans_parallel_informal.(v1.0)',
            data_root=_DATASET_ROOT['ko_en_trans_parallel'],
            feature=_TRANSLATION_KO_EN_PARALLEL_INFORMAL,
            data_sp_path={datasets.Split.TRAIN: ['1_구어체*.xlsx'],
                          datasets.Split.VALIDATION: ['1_구어체*.xlsx']},
            reading_fn=_parsing_ko_en_parallel_informal,
            parsing_fn=lambda x:x,
            split_fn=_DEFAULT_RAW_CORPUS_SPLIT,
        ),
        AIHubConfig(
            name='ko_en_trans_parallel_conversational.(v1.0)',
            data_root=_DATASET_ROOT['ko_en_trans_parallel'],
            feature=_TRANSLATION_KO_EN_PARALLEL_CONVERSATIONAL,
            data_sp_path={datasets.Split.TRAIN: ['2_대화체.xlsx'],
                          datasets.Split.VALIDATION: ['2_대화체.xlsx']},
            reading_fn=_parsing_ko_en_parallel_conversational,
            parsing_fn=lambda x:x,
            split_fn=_DEFAULT_RAW_CORPUS_SPLIT,
        ),
        AIHubConfig(
            name='ko_en_trans_parallel_news.(v1.0)',
            data_root=_DATASET_ROOT['ko_en_trans_parallel'],
            feature=_TRANSLATION_KO_EN_PARALLEL_NEWS,
            data_sp_path={datasets.Split.TRAIN: ['3_문어체_뉴스*.xlsx'],
                          datasets.Split.VALIDATION: ['3_문어체_뉴스*.xlsx']},
            reading_fn=_parsing_ko_en_parallel_news,
            parsing_fn=lambda x:x,
            split_fn=_DEFAULT_RAW_CORPUS_SPLIT,
        ),
        AIHubConfig(
            name='ko_en_trans_parallel_culture.(v1.0)',
            data_root=_DATASET_ROOT['ko_en_trans_parallel'],
            feature=_TRANSLATION_KO_EN_PARALLEL_CULTURE,
            data_sp_path={datasets.Split.TRAIN: ['4_문어체_한국문화.xlsx'],
                          datasets.Split.VALIDATION: ['4_문어체_한국문화.xlsx']},
            reading_fn=_parsing_ko_en_parallel_culture,
            parsing_fn=lambda x:x,
            split_fn=_DEFAULT_RAW_CORPUS_SPLIT,
        ),
        AIHubConfig(
            name='ko_en_trans_parallel_ordinance.(v1.0)',
            data_root=_DATASET_ROOT['ko_en_trans_parallel'],
            feature=_TRANSLATION_KO_EN_PARALLEL_ORDINANCE_WEB,
            data_sp_path={datasets.Split.TRAIN: ['5_문어체_조례.xlsx'],
                          datasets.Split.VALIDATION: ['5_문어체_조례.xlsx']},
            reading_fn=_parsing_ko_en_parallel_ordinance_web,
            parsing_fn=lambda x:x,
            split_fn=_DEFAULT_RAW_CORPUS_SPLIT,
        ),
        AIHubConfig(
            name='ko_en_trans_parallel_web.(v1.0)',
            data_root=_DATASET_ROOT['ko_en_trans_parallel'],
            feature=_TRANSLATION_KO_EN_PARALLEL_ORDINANCE_WEB,
            data_sp_path={datasets.Split.TRAIN: ['6_문어체_지자체웹사이트.xlsx'],
                          datasets.Split.VALIDATION: ['6_문어체_지자체웹사이트.xlsx']},
            reading_fn=_parsing_ko_en_parallel_ordinance_web,
            parsing_fn=lambda x:x,
            split_fn=_DEFAULT_RAW_CORPUS_SPLIT,
        ),

        AIHubConfig(
            name='translation_ko_ja.v1.0',
            data_root=_DATASET_ROOT['ko_ja_trans'],
            feature=_TRANSLATION_KO_JA_FEATURE,
            data_sp_path={datasets.Split.TRAIN: ['Training/*.json'],
                          datasets.Split.VALIDATION: ['Validation/*.json']},
            reading_fn=_parsing_specialty_ko_ja,
            parsing_fn=lambda x:x,
        ),
    
        AIHubConfig(
            name='translation_ko_zh_tech.v1.0',
            data_root=_DATASET_ROOT['ko_zh_trans_tech'],
            feature=_TRANSLATION_KO_ZH_FEATURE,
            data_sp_path={datasets.Split.TRAIN: ['Training/*.json'],
                          datasets.Split.VALIDATION: ['Validation/*.json']},
            reading_fn=_parsing_specialty_ko_zh,
            parsing_fn=lambda x:x,
        ),

        AIHubConfig(
            name='translation_ko_zh_social.v1.0',
            data_root=_DATASET_ROOT['ko_zh_trans_social'],
            feature=_TRANSLATION_KO_ZH_FEATURE,
            data_sp_path={datasets.Split.TRAIN: ['Training/*.json'],
                          datasets.Split.VALIDATION: ['Validation/*.json']},
            reading_fn=_parsing_specialty_ko_zh,
            parsing_fn=lambda x:x,
        ),
    ]

    MANUAL_DOWNLOAD_INSTRUCTIONS = """
    For the NIKL, you must manually download NIKL data from https://aihub.or.kr/
    and extract it under the proper location.
    all the data have to located under manual_dir/AIHub.
    This is dataset and path pairs. (all the paths are case-sensitive!)
    ============================================
    specialty_corpus.paper.(v1.0): manual_dir/AIHub/전문분야 말뭉치/Training/논문*.json
                                   manual_dir/AIHub/전문분야 말뭉치/Validation/논문*.json
                                   
    specialty_corpus.statute.(v1.0): manual_dir/AIHub/전문분야 말뭉치/Training/법령*.json
                                     manual_dir/AIHub/전문분야 말뭉치/Validation/법령*.json

    specialty_corpus.patent_n.(v1.0): manual_dir/AIHub/전문분야 말뭉치/Training/Training/특허_[0-9].json
                                      manual_dir/AIHub/전문분야 말뭉치/Validation/특허_[0-9].json

    specialty_corpus.patent_a.(v1.0): manual_dir/AIHub/전문분야 말뭉치/Training/특허_z*.json
                                      manual_dir/AIHub/전문분야 말뭉치/Validation/특허_z*.json

    specialty_corpus.leading_case.(v1.0): manual_dir/AIHub/전문분야 말뭉치/Training/판례*.json
                                          manual_dir/AIHub/전문분야 말뭉치/Validation/판례*.json

    specialty_ko_en.(v1.0): manual_dir/AIHub/전문분야한영/Training/*.json
                            manual_dir/AIHub/전문분야한영/Validation/*.json  

    korean_sns.(v1.0): manual_dir/AIHub/한국어 SNS/Training/*.json
                       manual_dir/AIHub/한국어 SNS/Validation/*.json

    korean_dialog.(v1.0): manual_dir/AIHub/한국어 대화/*.xlsx

    korean_dialog_summary.(v1.0): manual_dir/AIHub/한국어 대화 요약/Training/*.json
                                  manual_dir/AIHub/한국어 대화 요약/Validation/*.json  

    translation_ko_en_tech.(v1.0): manual_dir/AIHub/한국어-영어 번역 말뭉치(기술과학)/Training/*.json
                                   manual_dir/AIHub/한국어-영어 번역 말뭉치(기술과학)/Validation/*.json

    translation_ko_en_social.(v1.0): manual_dir/AIHub/한국어-영어 번역 말뭉치(사회과학)/Training/*.json
                                     manual_dir/AIHub/한국어-영어 번역 말뭉치(사회과학)/Validation/*.json

    ko_en_trans_parallel_informal.(v1.0): manual_dir/AIHub/한국어-영어 번역(병렬) 말뭉치/1_구어체*.xlsx
    ko_en_trans_parallel_informal.(v1.0): manual_dir/AIHub/한국어-영어 번역(병렬) 말뭉치/2_대화체.xlsx
    ko_en_trans_parallel_informal.(v1.0): manual_dir/AIHub/한국어-영어 번역(병렬) 말뭉치/3_문어체_뉴스*.xlsx
    ko_en_trans_parallel_informal.(v1.0): manual_dir/AIHub/한국어-영어 번역(병렬) 말뭉치/4_문어체_한국문화.xlsx
    ko_en_trans_parallel_informal.(v1.0): manual_dir/AIHub/한국어-영어 번역(병렬) 말뭉치/5_문어체_조례.xlsx
    ko_en_trans_parallel_informal.(v1.0): manual_dir/AIHub/한국어-영어 번역(병렬) 말뭉치/6_문어체_지자체웹사이트.xlsx

    translation_ko_ja.(v1.0): manual_dir/AIHub/한국어-일본어 번역 말뭉치/Training/*.json
                              manual_dir/AIHub/한국어-일본어 번역 말뭉치/Validation/*.json

    translation_ko_zh_tech.(v1.0): manual_dir/AIHub/AIHub/한국어-중국어 번역 말뭉치(기술과학)/Training/*.json
                                   manual_dir/AIHub/AIHub/한국어-중국어 번역 말뭉치(기술과학)/Validation/*.json

    translation_ko_zh_tech.(v1.0): manual_dir/AIHub/AIHub/한국어-중국어 번역 말뭉치(사회과학)/Training/*.json
                                   manual_dir/AIHub/AIHub/한국어-중국어 번역 말뭉치(사회과학)/Validation/*.json
                                                        
    ============================================
    """

    def _info(self) -> datasets.DatasetInfo:
        """Returns the dataset metadata."""
        return datasets.DatasetInfo(
            description=_DESCRIPTION,
            features=self.config.feature,
            homepage=self.config.homepage,
            citation=_CITATION,
        )

    def _split_generators(self, dl_manager: datasets.DownloadManager):
        """Returns SplitGenerators."""
        path_kv = {}
        for k, v in self.config.data_sp_path.items():
            path_list = []
            for vv in v:
                path_list.extend(glob.glob(os.path.join(
                    dl_manager.manual_dir, self.config.data_root, vv)))
            path_kv[k] = path_list
            path_list = []

        for _, v in path_kv.items():
            if len(v) == 0:
                raise AssertionError("For the AIHub dataset, you must manually download and extract dataset under {0}/{1}.".format(
                    dl_manager.manual_dir,
                    self.config.data_root
                ))

        if self.config.split_fn is not None:
            in_files = []
            for sp_s_key in self.config.split_fn['source']:
                in_files.extend(path_kv[sp_s_key])
            split_fn_kv = self.config.split_fn['split']
            return [
                datasets.SplitGenerator(name=k, gen_kwargs={'path_list': in_files, 'split_fn': v}) for k, v in split_fn_kv.items()
            ]

        return [
                datasets.SplitGenerator(name=k, gen_kwargs={'path_list': v}) for k, v in path_kv.items()
        ]

    def _generate_examples(self, path_list, split_fn=None):
        """Yields examples."""
        if split_fn is not None:
            split_filter = functools.partial(_filter_fn_hash_id, split_fn=split_fn)

        _hash_set = set()
        for file_path in path_list:
            try:
                for example in iter(self.config.reading_fn(file_path)):
                    uid, ex = self.config.parsing_fn(example)
                    
                    if split_fn is not None:
                        if not split_filter(str(uid)):
                            continue
                    hash_id = _hash_text(str(uid))
                    if hash_id not in _hash_set:
                        _hash_set.add(hash_id)
                        yield uid, ex
            except Exception as e:
                print(e)
